# SPDX-License-Identifier: Apache-2.0
"""Stage 1-2: crawl the Hugging Face Hub for SO-101/SO-100 datasets and screen them.

Searches the Hub dataset index by name, fetches `meta/info.json` and `meta/tasks.jsonl`
for every hit, and writes a catalog with a coarse `category` per repository.

category: clean / sim / blacklist_cam / small (episodes < 50) / parse_fail
  - blacklist_cam: laptop/phone/webcam/screen/iphone/android viewpoints only
  - depth cameras are kept and flagged with has_depth, not excluded

Input:  the Hub dataset index (network), or an existing catalog CSV in re-process mode.
Output: 02_so101_catalog.csv and 02_so101_catalog.xlsx (sheets: catalog, recommended).

Run:
  python 01_hf_so101_crawler.py --full       # full crawl
  python 01_hf_so101_crawler.py              # re-fetch metadata for rows already in the CSV
  python 01_hf_so101_crawler.py --rebuild    # rebuild the xlsx from the CSV, no network
"""

import argparse
import csv
import json
import re
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor

SEARCH_TERMS = ["so101", "so-101", "so100", "so-100",
                "soarm101", "soarm100", "so_arm101", "so_arm100", "soarm"]
SIM = re.compile(r"sim|isaac|mujoco|genesis|svla|synthetic", re.I)
VARIANT = re.compile(r"merged|multiplied|trimmed|converted|_copy|backup|eval_", re.I)
CAM_BLACKLIST = ("laptop", "phone", "webcam", "screen", "iphone", "android")
CAM_WHITELIST = ("wrist", "top", "front", "side", "overhead", "base", "ego",
                 "gripper", "hand", "head", "realsense", "depth", "third", "bird")
USER_AGENT = "so101-catalog-crawler"

COLUMNS = ["category", "id", "url", "codebase_version", "downloads", "likes",
           "last_modified", "robot_type", "action_dim", "arm", "n_cam", "cam_keys",
           "has_depth", "episodes", "frames", "fps", "is_variant", "score", "tasks", "tags"]


def metas_full() -> list:
    from huggingface_hub import HfApi
    api, repos = HfApi(), {}
    for q in SEARCH_TERMS:
        try:
            for d in api.list_datasets(search=q):
                if d.id in repos:
                    continue
                lm = getattr(d, "last_modified", None)
                repos[d.id] = {"id": d.id, "downloads": getattr(d, "downloads", 0) or 0,
                               "likes": getattr(d, "likes", 0) or 0,
                               "last_modified": str(lm)[:10] if lm else "",
                               "tags": ",".join(getattr(d, "tags", []) or [])}
        except Exception as e:
            print(f"  [list failed] {q}: {e}", flush=True)
        print(f"  '{q}' cumulative {len(repos)}", flush=True)
    return list(repos.values())


def metas_from_csv(csv_path: str) -> list:
    out = []
    for r in csv.DictReader(open(csv_path, encoding="utf-8")):
        out.append({"id": r["id"], "downloads": int(r["downloads"] or 0),
                    "likes": int(r["likes"] or 0), "last_modified": r["last_modified"],
                    "tags": r["tags"]})
    return out


def _get(url: str, tries: int = 4) -> bytes | None:
    for i in range(tries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
            with urllib.request.urlopen(req, timeout=15) as r:
                return r.read()
        except urllib.error.HTTPError as e:
            if e.code == 404:
                return None  # really absent, retrying is pointless
            if e.code in (429, 500, 503) and i < tries - 1:
                time.sleep(1.5 * (i + 1))
                continue
            return None
        except (urllib.error.URLError, TimeoutError, ConnectionError):
            if i < tries - 1:
                time.sleep(1.0 * (i + 1))
                continue
            return None
    return None


def fetch_meta(repo_id: str) -> tuple:
    base = f"https://huggingface.co/datasets/{repo_id}/resolve/main/meta"
    info = None
    raw = _get(f"{base}/info.json")
    if raw:
        try:
            info = json.loads(raw.decode("utf-8"))
        except json.JSONDecodeError:
            info = None
    tasks = []
    raw_t = _get(f"{base}/tasks.jsonl")  # v3 stores tasks as parquet, so this may 404
    if raw_t:
        for line in raw_t.decode("utf-8", "ignore").splitlines():
            try:
                t = json.loads(line).get("task")
                if t:
                    tasks.append(str(t))
            except json.JSONDecodeError:
                pass
    return info, list(dict.fromkeys(tasks))[:8]


def analyze(meta: dict, info, tasks: list) -> dict:
    row = {c: "" for c in COLUMNS}
    row.update({"id": meta["id"], "url": f"https://huggingface.co/datasets/{meta['id']}",
                "downloads": meta["downloads"], "likes": meta["likes"],
                "last_modified": meta["last_modified"], "tags": meta["tags"][:120],
                "is_variant": "Y" if VARIANT.search(meta["id"]) else "",
                "tasks": (" | ".join(tasks))[:240]})
    is_sim = bool(SIM.search(meta["id"]) or SIM.search(meta["tags"]))

    if info is None:
        row.update({"category": "parse_fail", "score": 0})
        return row

    robot = str(info.get("robot_type") or "").lower()
    feats = info.get("features", {}) or {}
    action = feats.get("action", {}) or {}
    shape = action.get("shape") or [0]
    adim = int(shape[0]) if shape else 0
    names = " ".join(map(str, action.get("names") or [])).lower()
    cams = [k.split(".")[-1] for k in feats if k.startswith("observation.images")]
    cam_lc = " ".join(cams).lower()
    has_black = any(b in cam_lc for b in CAM_BLACKLIST)
    has_depth = "depth" in cam_lc
    n_white = sum(1 for k in cams if any(w in k.lower() for w in CAM_WHITELIST))
    eps = int(info.get("total_episodes") or 0)

    row.update({"codebase_version": info.get("codebase_version") or "",
                "robot_type": robot or "(empty)", "action_dim": adim,
                "arm": "bi" if (adim >= 12 or ("left" in names and "right" in names)) else "single",
                "n_cam": len(cams), "cam_keys": ",".join(cams),
                "has_depth": "Y" if has_depth else "", "episodes": eps,
                "frames": int(info.get("total_frames") or 0), "fps": info.get("fps") or 0})

    score = min(eps, 300) + n_white * 20 + min(meta["downloads"], 400) // 4
    score -= 30 if has_black else 0
    score -= 15 if row["is_variant"] else 0
    row["score"] = score

    row["category"] = ("sim" if is_sim else "blacklist_cam" if has_black
                       else "small" if eps < 50 else "clean")
    return row


def write_xlsx(rows: list, xlsx_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hf, hfill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="4472C4")
    lf, cfill = Font(color="0563C1", underline="single"), PatternFill("solid", fgColor="E2EFDA")
    dfill = PatternFill("solid", fgColor="FFF2CC")  # depth highlight

    def fill(ws, data):
        ws.append(COLUMNS)
        for i in range(1, len(COLUMNS) + 1):
            ws.cell(1, i).font, ws.cell(1, i).fill = hf, hfill
            ws.cell(1, i).alignment = Alignment(horizontal="center")
        for r in data:
            ws.append([r.get(c, "") for c in COLUMNS])
            ri = ws.max_row
            uc = ws.cell(ri, COLUMNS.index("url") + 1)
            uc.hyperlink, uc.value, uc.font = r["url"], "open", lf
            if r.get("category") == "clean":
                ws.cell(ri, 1).fill = cfill
            if r.get("has_depth") == "Y":
                ws.cell(ri, COLUMNS.index("has_depth") + 1).fill = dfill
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
        widths = {"id": 46, "url": 6, "cam_keys": 26, "tasks": 50, "tags": 34,
                  "last_modified": 12, "codebase_version": 9}
        for i, c in enumerate(COLUMNS, 1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 10)

    ws1 = wb.active
    ws1.title = "catalog"
    fill(ws1, sorted(rows, key=lambda r: r.get("score", 0), reverse=True))
    rec = sorted([r for r in rows if r.get("category") == "clean"],
                 key=lambda r: (int(r["downloads"]), r.get("score", 0)), reverse=True)
    fill(wb.create_sheet("recommended"), rec)
    wb.save(xlsx_path)


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--csv", default="02_so101_catalog.csv", help="catalog CSV path")
    ap.add_argument("--xlsx", default="02_so101_catalog.xlsx", help="catalog XLSX path")
    ap.add_argument("--full", action="store_true", help="crawl the Hub index from scratch")
    ap.add_argument("--rebuild", action="store_true",
                    help="rebuild the XLSX from the CSV without any network access")
    ap.add_argument("--workers", type=int, default=12, help="metadata fetch workers")
    args = ap.parse_args()

    if args.rebuild:
        keep = [r for r in csv.DictReader(open(args.csv, encoding="utf-8"))
                if r["category"] in ("clean", "sim") and int(r["episodes"] or 0) >= 50]
        for r in keep:
            for k in ("downloads", "likes", "episodes", "frames", "score"):
                r[k] = int(r[k] or 0)
        write_xlsx(keep, args.xlsx)
        cat = {}
        for r in keep:
            cat[r["category"]] = cat.get(r["category"], 0) + 1
        print(f"[rebuild] xlsx updated, {len(keep)} rows "
              f"(blacklist/small/parse_fail excluded, episodes >= 50) {cat}", flush=True)
        return

    print(f"[1/4] collecting metadata ({'full crawl' if args.full else 'CSV re-process'})...",
          flush=True)
    metas = metas_full() if args.full else metas_from_csv(args.csv)
    print(f"  {len(metas)} repositories\n[2/4] fetching info+tasks (workers {args.workers})...",
          flush=True)

    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        results = list(ex.map(lambda m: fetch_meta(m["id"]), metas))
    print(f"  fetch done ({sum(i is not None for i, _ in results)} info / "
          f"{sum(1 for _, t in results if t)} with tasks)", flush=True)

    print("[3/4] analyzing...", flush=True)
    rows = [analyze(m, info, tasks) for m, (info, tasks) in zip(metas, results)]

    print("[4/4] writing...", flush=True)
    with open(args.csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows)
    write_xlsx(rows, args.xlsx)

    cat = {}
    for r in rows:
        cat[r["category"]] = cat.get(r["category"], 0) + 1
    clean = [r for r in rows if r["category"] == "clean"]
    cb = {}
    for r in clean:
        cb[r["codebase_version"] or "?"] = cb.get(r["codebase_version"] or "?", 0) + 1
    print(f"\n  category: {cat}")
    print(f"  clean {len(clean)} / episodes>=100 {sum(1 for r in clean if r['episodes']>=100)} "
          f"/ depth {sum(1 for r in clean if r['has_depth']=='Y')} / with tasks "
          f"{sum(1 for r in clean if r['tasks'])}")
    print(f"  clean codebase_version: {cb}")


if __name__ == "__main__":
    main()
