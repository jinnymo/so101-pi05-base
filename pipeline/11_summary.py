# SPDX-License-Identifier: Apache-2.0
"""Stage 6f: aggregate the kept pool into a `summary` sheet at the front of the report.

Four views: per class, per camera viewpoint (a dataset can appear in several rows),
fps / resolution / codec distribution, and the headline totals.

Input:  the stage 6a XLSX report, after stage 6c has written the `class` column.
Output: the same XLSX with a `summary` sheet inserted as the first sheet.

Run:
  python 11_summary.py --report so101_external_analysis.xlsx
"""

import argparse
from collections import Counter

TOP = ["top", "overhead", "bird", "high"]
WRIST = ["wrist", "handeye", "gripper", "endeff", "tip", "eye"]
FRONT = ["front"]
SIDE = ["side", "left", "right", "lateral"]


def num(x):
    try:
        return int(float(x))
    except (ValueError, TypeError):
        return 0


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--report", default="so101_external_analysis.xlsx", help="stage 6a XLSX")
    args = ap.parse_args()

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    wb = load_workbook(args.report)
    ws = wb["datasets"]
    H = [c.value for c in ws[1]]
    R = [dict(zip(H, [c.value for c in row])) for row in ws.iter_rows(min_row=2)]
    conf = [r for r in R if str(r["class"]).startswith(("1", "2", "3"))]
    excl = [r for r in R if str(r["class"]).startswith("9")]

    if "summary" in wb.sheetnames:
        del wb["summary"]
    s = wb.create_sheet("summary", 0)
    bold = Font(bold=True)
    hd = Font(bold=True, color="FFFFFF")
    hf = PatternFill("solid", fgColor="4472C4")
    grn = PatternFill("solid", fgColor="C6EFCE")
    row = [1]

    def put(vals, style=None, fill=None):
        for j, v in enumerate(vals, 1):
            c = s.cell(row[0], j, v)
            if style:
                c.font = style
            if fill:
                c.fill = fill
        row[0] += 1

    def blank():
        row[0] += 1

    put(["SO-101 external pool summary", "", "", ""], Font(bold=True, size=14))
    blank()

    put(["[1] by class", "datasets", "episodes", "frames"], hd, hf)
    cls_order = [("1_keep_tier1", "tier1 (top + wrist)"), ("2_keep_tier2", "tier2 (one of them)"),
                 ("2b_keep_tier2b", "tier2b (visually confirmed)"),
                 ("3_keep_tier3", "tier3 (front / side)")]
    tot_n = tot_e = tot_f = 0
    for cls, label in cls_order:
        rs = [r for r in R if r["class"] == cls]
        e = sum(num(r["episodes"]) for r in rs)
        f = sum(num(r["frames"]) for r in rs)
        put([label, len(rs), e, f])
        tot_n += len(rs)
        tot_e += e
        tot_f += f
    put(["kept total", tot_n, tot_e, tot_f], bold, grn)
    ex_e = sum(num(r["episodes"]) for r in excl)
    put(["(reference) excluded", len(excl), ex_e, sum(num(r["frames"]) for r in excl)])
    blank()

    put(["[2] camera viewpoint (kept, overlapping)", "datasets", "episodes", ""], hd, hf)

    def has(r, kws):
        cl = str(r["cam_keys"]).lower()
        return any(k in cl for k in kws)
    for label, kws in [("top", TOP), ("wrist", WRIST), ("front", FRONT), ("side", SIDE)]:
        rs = [r for r in conf if has(r, kws)]
        put([label, len(rs), sum(num(r["episodes"]) for r in rs), ""])
    blank()

    put(["[3] fps / resolution / codec (kept)", "datasets", "", ""], hd, hf)
    for label, key in [("fps", "fps"), ("resolution", "resolution"), ("codec", "codec")]:
        cnt = Counter(str(r[key]) for r in conf)
        put([label, "", "", ""], bold)
        for k, v in cnt.most_common(6):
            put([f"   {k}", v, "", ""])
    blank()

    put(["[4] totals", "", "", ""], hd, hf)
    put(["kept datasets", tot_n, "", ""])
    put(["episodes", tot_e, "", ""], bold, grn)
    put(["frames", tot_f, "", ""], bold, grn)
    put(["mean episode length (frames)", round(tot_f / tot_e, 1) if tot_e else 0, "", ""])
    put(["estimated video hours (at 30 fps)", round(tot_f / 30 / 3600, 1), "", ""])

    for col, w in {"A": 30, "B": 11, "C": 11, "D": 11}.items():
        s.column_dimensions[col].width = w
    s.freeze_panes = "A2"
    wb.save(args.report)

    print(f"kept {tot_n} datasets / {tot_e} episodes / {tot_f} frames")
    if tot_e:
        print(f"  {tot_f/tot_e:.0f} frames per episode / about {tot_f/30/3600:.1f} hours at 30 fps")
    print("  summary sheet inserted as the first sheet")


if __name__ == "__main__":
    main()
