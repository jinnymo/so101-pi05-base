# SPDX-License-Identifier: Apache-2.0
"""Stage 6g: list kept datasets whose task string is empty, so prompts can be filled in.

A dataset with no task string cannot be used as a language-conditioned sample. This writes
a worksheet with one row per affected dataset, a hint extracted from the repository name,
a Hub link, and an empty column to write the real prompt into.

Input:  the stage 6a XLSX report, after stage 6c has written the `class` column.
Output: the same XLSX with a `prompt_missing` sheet inserted as the second sheet.

Run:
  python 12_prompt_none.py --report so101_external_analysis.xlsx
"""

import argparse
import re


def hint(d):
    rest = d.split("__", 1)[-1]
    rest = re.sub(r"so-?arm-?\d*|so-?101|so-?100", "", rest, flags=re.I)
    rest = re.sub(r"v?\d+(ep|eps|fps|sec|hz|s|d|loc)?\b", "", rest, flags=re.I)
    rest = re.sub(r"\b(test|eval|merged|train|trimmed|final|dataset|record|new|fixed|"
                  r"converted|repaired|cropped|copy|appendix|v\d)\b", "", rest, flags=re.I)
    rest = re.sub(r"[_\-]+", " ", rest)
    rest = re.sub(r"\d{6,}", "", rest)
    return re.sub(r"\s+", " ", rest).strip()


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--report", default="so101_external_analysis.xlsx", help="stage 6a XLSX")
    args = ap.parse_args()

    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(args.report)
    ws = wb["datasets"]
    H = [c.value for c in ws[1]]
    R = [dict(zip(H, [c.value for c in row])) for row in ws.iter_rows(min_row=2)]
    none_t = [r for r in R
              if str(r["class"]).startswith(("1", "2", "3"))
              and (not str(r["tasks"]).strip() or str(r["tasks"]).strip() == "None")]
    none_t.sort(key=lambda r: (r["class"], r["dir"]))

    name = "prompt_missing"
    if name in wb.sheetnames:
        del wb[name]
    s = wb.create_sheet(name, 1)
    cols = ["class", "dir", "url", "episodes", "cam_keys", "name_hint", "task_to_fill"]
    s.append(cols)
    hd, hf = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="C55A11")
    lf = Font(color="0563C1", underline="single")
    yel = PatternFill("solid", fgColor="FFF2CC")
    for i in range(1, len(cols) + 1):
        s.cell(1, i).font, s.cell(1, i).fill = hd, hf
        s.cell(1, i).alignment = Alignment(horizontal="center")
    for r in none_t:
        s.append([r["class"], r["dir"], "HF", r["episodes"], r["cam_keys"], hint(r["dir"]), ""])
        ri = s.max_row
        uc = s.cell(ri, 3)
        uc.hyperlink, uc.font = f"https://huggingface.co/datasets/{r['dir'].replace('__', '/', 1)}", lf
        s.cell(ri, 7).fill = yel
    s.freeze_panes = "A2"
    s.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{s.max_row}"
    for col, w in {"A": 16, "B": 46, "C": 5, "D": 8, "E": 24, "F": 30, "G": 40}.items():
        s.column_dimensions[col].width = w
    wb.save(args.report)
    print(f"{len(none_t)} datasets without a task string -> 'prompt_missing' sheet")


if __name__ == "__main__":
    main()
