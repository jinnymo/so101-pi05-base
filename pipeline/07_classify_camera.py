# SPDX-License-Identifier: Apache-2.0
"""Stage 6b: classify datasets by camera viewpoint and move them into tier directories.

  confirmed/tier12  datasets that carry a top or a wrist view (the target layout)
  confirmed/tier3   no top and no wrist, but an unambiguous front or side view; kept only
                    up to a quota of 25% of the tier12 episode count, best-scoring first
  undecided         "up" views and unclear keys (camera_N, main, ego, ...): needs a look
  excluded          tier3 datasets beyond the quota

The quota keeps the pool dominated by viewpoints that match the target robot instead of
letting generic front-facing recordings take over.

Input:  the stage 6a XLSX report and <root>/external_hf/*.
Output: the same datasets, moved into the four subdirectories above.

Run:
  python 07_classify_camera.py --root /path/to/workspace --report so101_external_analysis.xlsx
"""

import argparse
import os
import shutil

TOP = ["top", "overhead", "bird", "high"]
WRIST = ["wrist", "handeye", "gripper", "endeff", "tip", "eye"]
FRONTSIDE = ["front", "side", "left", "right", "lateral"]
RATIO_TIER3 = 0.25  # tier3 episodes <= tier12 episodes * 0.25


def has(cl, kws):
    return any(k in cl for k in kws)


def epn(r):
    try:
        return int(float(r["episodes"]))
    except (ValueError, TypeError):
        return 0


def anom(r, c):
    try:
        return int(str(r[c]).split("/")[0])
    except (ValueError, AttributeError):
        return 0


def score(r):
    s = min(epn(r), 300) + min(int(float(r["downloads"] or 0)), 400) // 4
    s -= anom(r, "n_anomaly_ep") * 10 + anom(r, "n_static_ep") * 5
    s += int(r["n_cam"] or 0) * 10
    return s


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--root", default=os.environ.get("ROOT", "."),
                    help="workspace root (default: $ROOT)")
    ap.add_argument("--src", default=None, help="dataset directory (default: <root>/external_hf)")
    ap.add_argument("--report", default="so101_external_analysis.xlsx", help="stage 6a XLSX")
    args = ap.parse_args()

    from openpyxl import load_workbook

    src = args.src or os.path.join(args.root, "external_hf")
    ws = load_workbook(args.report)["datasets"]
    H = [c.value for c in ws[1]]
    R = [dict(zip(H, [c.value for c in row])) for row in ws.iter_rows(min_row=2)]

    tier12, tier3, undecided = [], [], []
    for r in R:
        cl = str(r["cam_keys"]).lower()
        if has(cl, TOP) or has(cl, WRIST):
            tier12.append(r)
        elif "up" in cl:
            undecided.append(r)
        elif has(cl, FRONTSIDE):
            tier3.append(r)
        else:
            undecided.append(r)

    t12_ep = sum(epn(r) for r in tier12)
    budget = t12_ep * RATIO_TIER3
    tier3.sort(key=score, reverse=True)
    keep, excl, acc = [], [], 0
    for r in tier3:
        if acc + epn(r) <= budget:
            keep.append(r)
            acc += epn(r)
        else:
            excl.append(r)

    def move(rows, sub):
        os.makedirs(f"{src}/{sub}", exist_ok=True)
        n = 0
        for r in rows:
            path = f"{src}/{r['dir']}"
            if os.path.isdir(path):
                shutil.move(path, f"{src}/{sub}/{r['dir']}")
                n += 1
        return n

    n12 = move(tier12, "confirmed/tier12")
    n3 = move(keep, "confirmed/tier3")
    nu = move(undecided, "undecided")
    ne = move(excl, "excluded")

    print(f"confirmed/tier12 (top or wrist): {n12} datasets / {t12_ep} episodes")
    print(f"confirmed/tier3  (best front or side): {n3} datasets / {acc} episodes "
          f"(quota <= {budget:.0f})")
    print(f"undecided (up or unclear): {nu} datasets")
    print(f"excluded (over the tier3 quota): {ne} datasets / {sum(epn(r) for r in excl)} episodes")
    total_keep = t12_ep + acc
    if total_keep:
        print(f"\nkept: {n12 + n3} datasets / {total_keep} episodes "
              f"(top+wrist {t12_ep/total_keep*100:.0f}% : front+side {acc/total_keep*100:.0f}%)")


if __name__ == "__main__":
    main()
