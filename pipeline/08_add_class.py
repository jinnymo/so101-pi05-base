# SPDX-License-Identifier: Apache-2.0
"""Stage 6c: add a `class` column to the report from the directory each dataset sits in.

No re-analysis: the class is read from directory membership, so whatever manual sorting
happened after stage 6b is what ends up in the spreadsheet. Rows are then sorted by class
and colour-coded.

Between stage 6b and this stage the tier12 bucket is split by hand into tier1 (top and
wrist) and tier2 (one of the two), and the datasets held back in `undecided` are reviewed
visually; the ones that pass move to `confirmed/tier2b`.

Input:  <root>/external_hf/{confirmed/tier1,confirmed/tier2,confirmed/tier2b,
        confirmed/tier3,excluded} and the stage 6a XLSX report.
Output: the same XLSX, with `class` and `is_sim` columns prepended.

Run:
  python 08_add_class.py --root /path/to/workspace --report so101_external_analysis.xlsx
"""

import argparse
import os
import re

CLSMAP = [("confirmed/tier1", "1_keep_tier1"), ("confirmed/tier2", "2_keep_tier2"),
          ("confirmed/tier2b", "2b_keep_tier2b"), ("confirmed/tier3", "3_keep_tier3"),
          ("excluded", "9_excluded")]
FILLS = {"1_keep_tier1": "C6EFCE", "2_keep_tier2": "D9EAD3", "2b_keep_tier2b": "D9EAD3",
         "3_keep_tier3": "E2EFDA", "9_excluded": "FFC7CE"}
SIM_RE = re.compile(r"isaac|leisaac|svla|mujoco|genesis|sim_so101|__sim", re.I)


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--root", default=os.environ.get("ROOT", "."),
                    help="workspace root (default: $ROOT)")
    ap.add_argument("--src", default=None, help="dataset directory (default: <root>/external_hf)")
    ap.add_argument("--report", default="so101_external_analysis.xlsx", help="stage 6a XLSX")
    args = ap.parse_args()

    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    src = args.src or os.path.join(args.root, "external_hf")
    loc = {}
    for sub, cls in CLSMAP:
        p = f"{src}/{sub}"
        if os.path.isdir(p):
            for n in os.listdir(p):
                loc[n] = cls

    ws = load_workbook(args.report)["datasets"]
    H = [c.value for c in ws[1]]
    rows = [dict(zip(H, [c.value for c in row])) for row in ws.iter_rows(min_row=2)]
    for r in rows:
        r["class"] = loc.get(r["dir"], "?")
        r["is_sim"] = "sim" if SIM_RE.search(str(r["dir"])) else ""
    cols = ["class", "is_sim"] + H
    order = {cls: i for i, (_, cls) in enumerate(CLSMAP)}
    rows.sort(key=lambda r: (order.get(r["class"], 9), -int(float(r.get("downloads") or 0))))

    wb = Workbook()
    ws2 = wb.active
    ws2.title = "datasets"
    hf, hfill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="4472C4")
    lf = Font(color="0563C1", underline="single")
    ws2.append(cols)
    for i in range(1, len(cols) + 1):
        ws2.cell(1, i).font, ws2.cell(1, i).fill = hf, hfill
        ws2.cell(1, i).alignment = Alignment(horizontal="center")
    ui = cols.index("url")
    for r in rows:
        ws2.append([r.get(c, "") for c in cols])
        ri = ws2.max_row
        ws2.cell(ri, 1).fill = PatternFill("solid", fgColor=FILLS.get(r["class"], "FFFFFF"))
        uc = ws2.cell(ri, ui + 1)
        if r.get("url"):
            uc.hyperlink, uc.value, uc.font = r["url"], "HF", lf
    ws2.freeze_panes = "C2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws2.max_row}"
    widths = {"class": 14, "dir": 42, "url": 5, "cam_keys": 20, "tasks": 36,
              "action_min": 26, "action_max": 26}
    for i, c in enumerate(cols, 1):
        ws2.column_dimensions[get_column_letter(i)].width = widths.get(c, 10)
    wb.save(args.report)

    cnt = {}
    for r in rows:
        cnt[r["class"]] = cnt.get(r["class"], 0) + 1
    print("class distribution:", dict(sorted(cnt.items())))


if __name__ == "__main__":
    main()
