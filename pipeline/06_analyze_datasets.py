# SPDX-License-Identifier: Apache-2.0
"""Stage 6a: per-dataset precision analysis of the downloaded pool.

Walks every directory under <root>/external_hf and reports, per dataset:
  - basics: codebase version, episodes, frames, fps, duration, size, downloads, tasks
  - cameras: count, keys, resolution, codec, depth presence
  - action: dimension, joint names, unit, per-joint min/max/mean, offset vs the reference
  - conformity: reference range overlap, action-state offset, stage 3 flag
  - trajectories: largest frame-to-frame action jump, static episodes, desynced episodes,
    per-joint range warnings, and whether the first video decodes to a non-black frame

Input:  <root>/external_hf/*, 03_action_match.csv, and the reference dataset (--ref).
Output: an XLSX report (--out), one row per dataset.

Run:
  python 06_analyze_datasets.py --root /path/to/workspace --ref /path/to/reference_dataset
"""

import argparse
import csv
import glob
import json
import os
import statistics
import threading
from concurrent.futures import ThreadPoolExecutor

COLUMNS = ["flag", "dir", "url", "downloads", "version", "episodes", "frames", "fps",
           "duration_s", "size_MB", "n_cam", "cam_keys", "resolution", "codec", "depth",
           "action_dim", "names_ok", "unit", "ref_overlap", "act_state_offset",
           "ref_0pt_offset", "traj_max_jump", "n_anomaly_ep", "n_static_ep",
           "n_desync_ep", "joint_range", "range_warn", "video",
           "action_min", "action_max", "tasks"]

JUMP_THRESH = 50.0    # frame-to-frame action change above this is an anomaly (degree)
STATIC_THRESH = 5.0   # every joint range below this over a whole episode: idle/empty (degree)
DESYNC_THRESH = 10.0  # mean |action.mean - state.mean| above this: action/state desync (degree)
RANGE_WIDE = 320.0    # joint range above this is abnormally wide (degree)
CAM_PRIORITY = ["front", "top", "overhead", "main", "base", "side", "third", "bird",
                "fixed", "high", "context", "handeye", "ego", "cam_top"]


def fmt(arr, p=0):
    return ",".join(f"{x:.{p}f}" for x in arr)


def load_action_stats(d):
    sp = os.path.join(d, "meta", "stats.json")
    if os.path.exists(sp):
        s = json.load(open(sp))
        if "action" in s:
            return s["action"], s.get("observation.state")
    ep = os.path.join(d, "meta", "episodes_stats.jsonl")
    if not os.path.exists(ep):
        return None, None
    amin = amax = amean = smean = None
    total = 0
    for line in open(ep, encoding="utf-8"):
        try:
            st = json.loads(line).get("stats", {})
        except json.JSONDecodeError:
            continue
        a = st.get("action")
        if not a:
            continue
        c = a["count"][0] if isinstance(a["count"], list) else a["count"]
        if amin is None:
            amin, amax = list(a["min"]), list(a["max"])
            amean = [m * c for m in a["mean"]]
        else:
            amin = [min(x, y) for x, y in zip(amin, a["min"])]
            amax = [max(x, y) for x, y in zip(amax, a["max"])]
            amean = [acc + m * c for acc, m in zip(amean, a["mean"])]
        s2 = st.get("observation.state")
        if s2:
            smean = ([m * c for m in s2["mean"]] if smean is None
                     else [acc + m * c for acc, m in zip(smean, s2["mean"])])
        total += c
    if amin is None or total == 0:
        return None, None
    act = {"min": amin, "max": amax, "mean": [x / total for x in amean]}
    state = {"mean": [x / total for x in smean]} if smean else None
    return act, state


def all_ep_full(d):
    """Scan action+state of every episode: jumps, idle episodes, desync, joint ranges."""
    import numpy as np
    import pyarrow.parquet as pq

    cands = sorted(glob.glob(os.path.join(d, "data", "chunk-*", "episode_*.parquet")))
    if not cands:
        cands = sorted(glob.glob(os.path.join(d, "data", "chunk-*", "*.parquet")))
    r = {"max_jump": 0.0, "n_anom": 0, "n_static": 0, "n_desync": 0,
         "n_ep": 0, "gmin": None, "gmax": None}
    for p in cands:
        try:
            tbl = pq.read_table(p, columns=["action", "observation.state"])
            has_state = True
        except Exception:
            try:
                tbl = pq.read_table(p, columns=["action"])
                has_state = False
            except Exception:
                continue
        act = np.array(tbl.column("action").to_pylist(), dtype=float)
        if act.ndim != 2 or len(act) < 2:
            continue
        r["n_ep"] += 1
        j = float(np.abs(np.diff(act, axis=0)).max())
        r["max_jump"] = max(r["max_jump"], j)
        if j > JUMP_THRESH:
            r["n_anom"] += 1
        if float((act.max(0) - act.min(0)).max()) < STATIC_THRESH:
            r["n_static"] += 1
        amn, amx = act.min(0), act.max(0)
        r["gmin"] = amn if r["gmin"] is None else np.minimum(r["gmin"], amn)
        r["gmax"] = amx if r["gmax"] is None else np.maximum(r["gmax"], amx)
        if has_state:
            try:
                st = np.array(tbl.column("observation.state").to_pylist(), dtype=float)
                if st.shape == act.shape and \
                        float(np.abs(act.mean(0) - st.mean(0)).mean()) > DESYNC_THRESH:
                    r["n_desync"] += 1
            except Exception:
                pass
    return r


def check_video(d, info):
    """Is the first episode of the representative camera a black screen?"""
    feats = info.get("features", {})
    vids = [k for k in feats if k.startswith("observation.images") and feats[k].get("dtype") == "video"]
    if not vids:
        return "no_video"
    cam = next((k for p in CAM_PRIORITY for k in vids if p in k.lower()), vids[0])
    paths = sorted(glob.glob(os.path.join(d, "videos", "chunk-*", cam, "episode_*.mp4")))
    if not paths:
        paths = sorted(glob.glob(os.path.join(d, "videos", cam, "chunk-*", "*.mp4")))
    if not paths:
        return "no_file"
    try:
        import av
        c = av.open(paths[0])
        means = []
        # First 10 frames: a single leading black frame is common and not a defect.
        for i, frame in enumerate(c.decode(video=0)):
            means.append(float(frame.to_ndarray(format="gray").mean()))
            if i >= 9:
                break
        c.close()
        return "BLACK" if (means and max(means) < 10) else "OK"
    except Exception as e:
        return f"err:{str(e)[:18]}"


def analyze_dir(d, match, ref_act):
    name = os.path.basename(d)
    info_path = os.path.join(d, "meta", "info.json")
    if not os.path.exists(info_path):
        return None
    info = json.load(open(info_path))
    feats = info.get("features", {})
    m = match.get(name, {})

    cams = [k for k in feats if k.startswith("observation.images")]
    cam_short = [k.split(".")[-1] for k in cams]
    res = codec = ""
    if cams:
        vi = feats[cams[0]].get("info", {}) or {}
        res = f"{vi.get('video.width','?')}x{vi.get('video.height','?')}"
        codec = vi.get("video.codec", "")
    eps = int(info.get("total_episodes") or 0)
    frames = int(info.get("total_frames") or 0)
    fps = info.get("fps") or 0

    act, state = load_action_stats(d)
    row = {c: "" for c in COLUMNS}
    row.update({
        "flag": m.get("flag", ""),
        "dir": name,
        "url": f"https://huggingface.co/datasets/{name.replace('__', '/', 1)}",
        "downloads": int(m.get("downloads") or 0),
        "version": info.get("codebase_version", ""),
        "episodes": eps, "frames": frames, "fps": fps,
        "duration_s": round(frames / fps, 1) if fps else "",
        "size_MB": round(sum(os.path.getsize(os.path.join(r, f))
                             for r, _, fs in os.walk(d) for f in fs) / 2**20),
        "n_cam": len(cams), "cam_keys": ",".join(cam_short),
        "resolution": res, "codec": codec,
        "depth": "Y" if "depth" in " ".join(cam_short).lower() else "",
        "action_dim": len(feats.get("action", {}).get("shape", [0])) and
        feats.get("action", {}).get("shape", [0])[0],
        "names_ok": m.get("names_ok", ""), "unit": m.get("unit", ""),
        "ref_overlap": m.get("range_overlap", ""),
        "act_state_offset": m.get("act_state_offset", ""),
        "tasks": (m.get("tasks") or "")[:120],
    })
    a = all_ep_full(d)
    ne = a["n_ep"]
    row["traj_max_jump"] = round(a["max_jump"] or 0, 1)
    row["n_anomaly_ep"] = f"{a['n_anom']}/{ne}"
    row["n_static_ep"] = f"{a['n_static']}/{ne}"
    row["n_desync_ep"] = f"{a['n_desync']}/{ne}"
    if a["gmin"] is not None:
        rng = a["gmax"] - a["gmin"]
        row["joint_range"] = ",".join(f"{x:.0f}" for x in rng)
        warn = []
        for i, rg in enumerate(rng):
            if rg > RANGE_WIDE:
                warn.append(f"j{i}:wide{rg:.0f}")
            elif rg < 1.0 and i < len(rng) - 1:  # near-frozen joint, gripper excluded
                warn.append(f"j{i}:stuck")
        row["range_warn"] = ";".join(warn)
    row["video"] = check_video(d, info)
    if act:
        row["action_min"] = fmt(act["min"])
        row["action_max"] = fmt(act["max"])
        # Task-independent approximation of a calibration offset.
        if ref_act and len(act["min"]) == len(ref_act["min"]):
            mids = [(act["min"][i] + act["max"][i]) / 2 for i in range(len(act["min"]))]
            refmids = [(ref_act["min"][i] + ref_act["max"][i]) / 2 for i in range(len(ref_act["min"]))]
            row["ref_0pt_offset"] = round(statistics.mean(abs(a - b) for a, b in zip(mids, refmids)), 1)
    return row


def write_xlsx(rows, out):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "datasets"
    hf, hfill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="4472C4")
    lf = Font(color="0563C1", underline="single")
    fills = {"green": PatternFill("solid", fgColor="C6EFCE"),
             "yellow": PatternFill("solid", fgColor="FFEB9C"),
             "red": PatternFill("solid", fgColor="FFC7CE")}
    order = {"green": 0, "yellow": 1, "red": 2, "": 3}
    rows.sort(key=lambda r: (order.get(r["flag"], 3), -int(r["downloads"] or 0)))

    ws.append(COLUMNS)
    for i in range(1, len(COLUMNS) + 1):
        ws.cell(1, i).font, ws.cell(1, i).fill = hf, hfill
        ws.cell(1, i).alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r.get(c, "") for c in COLUMNS])
        ri = ws.max_row
        ws.cell(ri, 1).fill = fills.get(r["flag"], PatternFill())
        uc = ws.cell(ri, COLUMNS.index("url") + 1)
        uc.hyperlink, uc.value, uc.font = r["url"], "HF", lf
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    widths = {"dir": 44, "url": 5, "cam_keys": 22, "tasks": 40, "action_min": 28,
              "action_max": 28, "resolution": 11, "version": 8}
    for i, c in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 10)
    wb.save(out)


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--root", default=os.environ.get("ROOT", "."),
                    help="workspace root (default: $ROOT)")
    ap.add_argument("--src", default=None, help="dataset directory (default: <root>/external_hf)")
    ap.add_argument("--ref", default=os.environ.get("REF", "reference_dataset"),
                    help="reference LeRobot dataset directory (default: $REF)")
    ap.add_argument("--match-csv", default="03_action_match.csv", help="stage 3 result CSV")
    ap.add_argument("--out", default="so101_external_analysis.xlsx", help="output XLSX")
    ap.add_argument("--workers", type=int, default=8)
    args = ap.parse_args()

    src = args.src or os.path.join(args.root, "external_hf")
    match = {r["id"].replace("/", "__"): r
             for r in csv.DictReader(open(args.match_csv, encoding="utf-8"))}
    ref_act = json.load(open(f"{args.ref}/meta/stats.json"))["action"]
    dirs = [d for d in sorted(glob.glob(f"{src}/*"))
            if os.path.isdir(d) and not d.endswith("_v3.0")]
    print(f"analyzing {len(dirs)} directories", flush=True)

    done = {"n": 0}
    lk = threading.Lock()

    def proc(d):
        try:
            r = analyze_dir(d, match, ref_act)
        except Exception as e:
            print(f"  failed {os.path.basename(d)}: {str(e)[:70]}", flush=True)
            r = None
        with lk:
            done["n"] += 1
            if done["n"] % 50 == 0:
                print(f"  {done['n']}/{len(dirs)}", flush=True)
        return r

    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        rows = [r for r in ex.map(proc, dirs) if r]

    write_xlsx(rows, args.out)
    print(f"\ndone: {len(rows)} rows -> {args.out}", flush=True)
    fl = {}
    for r in rows:
        fl[r["flag"] or "?"] = fl.get(r["flag"] or "?", 0) + 1
    print(f"  flag: {fl}")


if __name__ == "__main__":
    main()
