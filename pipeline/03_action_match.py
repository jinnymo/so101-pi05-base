# SPDX-License-Identifier: Apache-2.0
"""Stage 3: action-space conformity check, candidates vs a reference recording.

The reference defines the expected action space: joint names, unit (degree) and range,
recorded in the follower frame. For every candidate the script fetches `meta/info.json`
and `meta/stats.json` (falling back to `meta/episodes_stats.jsonl`) and compares
dimension, joint names and order, unit, action-vs-state offset and range overlap.

flag: green  (degree + names in the standard order + offset <= 6 deg)
      yellow (different unit / different order / offset 6-15 deg = weak leader suspicion)
      red    (dim != 6 / name mismatch / offset > 15 deg = strong leader suspicion / raw units)

A large mean gap between `action` and `observation.state` means the action column was
recorded in the leader's calibration frame, which injects a constant offset into training
targets. That is the single most damaging defect this stage catches.

Input:  02_so101_catalog.csv (rows with category clean or sim, episodes >= 50)
        and a reference LeRobot dataset directory (--ref).
Output: 03_action_match.csv and 03_action_match.xlsx.

Run:
  python 03_action_match.py --ref /path/to/reference_dataset
  python 03_action_match.py --retry     # re-run only rows whose fetch failed
  python 03_action_match.py --video     # attach a first-episode preview URL per row

If you have no SO-101 of your own, point --ref at a public dataset that this stage
would flag green; see pipeline/README.md.
"""

import argparse
import csv
import json
import os
import statistics
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor

STD = ["shoulder_pan", "shoulder_lift", "elbow_flex", "wrist_flex", "wrist_roll", "gripper"]
USER_AGENT = "so101-catalog-crawler"

COLUMNS = ["flag", "id", "url", "video_url", "downloads", "episodes", "codebase_version",
           "dim", "names_ok", "order_ok", "unit", "act_state_offset",
           "range_overlap", "robot_type", "action_min", "action_max", "tasks", "note"]

# Representative camera for the first-episode preview: a fixed external view tells you
# more about the scene and the objects than a wrist view does.
CAM_PRIORITY = ["front", "top", "overhead", "main", "base", "side", "third", "bird",
                "fixed", "high", "context", "handeye", "ego", "cam_top"]


def norm_name(n: str) -> str:
    return n.lower().replace(".pos", "").replace("_pos", "").replace("observation.", "").strip()


def classify_unit(amin: list, amax: list) -> str:
    m = max(max(abs(x) for x in amin), max(abs(x) for x in amax))
    if m < 1.5:
        return "norm"
    if m < 8:
        return "radian"
    if m < 400:
        return "degree"
    return "raw"


def load_ref(ref: str) -> dict:
    info = json.load(open(f"{ref}/meta/info.json"))
    stats = None
    if os.path.exists(f"{ref}/meta/stats.json"):
        stats = json.load(open(f"{ref}/meta/stats.json"))
    if not stats or "action" not in stats:
        # Most v2.1 datasets ship per-episode statistics only. Aggregate them the same
        # way candidates are aggregated, so any LeRobot dataset can serve as reference.
        stats = aggregate_episodes_stats(open(f"{ref}/meta/episodes_stats.jsonl", "rb").read())
    if not stats or "action" not in stats:
        raise SystemExit(f"{ref}: no action statistics in meta/stats.json "
                         f"or meta/episodes_stats.jsonl")
    names = [norm_name(n) for n in info["features"]["action"]["names"]]
    a = stats["action"]
    return {"names": names, "min": a["min"], "max": a["max"], "mean": a["mean"],
            "unit": classify_unit(a["min"], a["max"])}


def _get(url: str, tries: int = 4) -> bytes | None:
    for i in range(tries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
            with urllib.request.urlopen(req, timeout=15) as r:
                return r.read()
        except urllib.error.HTTPError as e:
            if e.code == 404:
                return None
            if e.code in (429, 500, 503) and i < tries - 1:
                time.sleep(1.5 * (i + 1))
                continue
            return None
        except (urllib.error.URLError, TimeoutError, ConnectionError):
            if i < tries - 1:
                time.sleep(1.0 * (i + 1))
                continue
            return None
    return None


def fetch_json(repo: str, fname: str):
    raw = _get(f"https://huggingface.co/datasets/{repo}/resolve/main/meta/{fname}")
    if not raw:
        return None
    try:
        return json.loads(raw.decode("utf-8"))
    except json.JSONDecodeError:
        return None


def aggregate_episodes_stats(raw: bytes):
    """Count-weighted aggregation of episodes_stats.jsonl, for datasets with no stats.json."""
    amin = amax = amean = smean = None
    total = 0
    for line in raw.decode("utf-8", "ignore").splitlines():
        try:
            d = json.loads(line).get("stats", {})
        except json.JSONDecodeError:
            continue
        a = d.get("action")
        if not a:
            continue
        c = a["count"][0] if isinstance(a["count"], list) else a["count"]
        if amin is None:
            amin, amax = list(a["min"]), list(a["max"])
            amean = [m * c for m in a["mean"]]
        else:
            amin = [min(x, y) for x, y in zip(amin, a["min"])]
            amax = [max(x, y) for x, y in zip(amax, a["max"])]
            amean = [acc + m * c for acc, m in zip(amean, a["mean"])]
        st = d.get("observation.state")
        if st:
            smean = ([m * c for m in st["mean"]] if smean is None
                     else [acc + m * c for acc, m in zip(smean, st["mean"])])
        total += c
    if amin is None or total == 0:
        return None
    out = {"action": {"min": amin, "max": amax, "mean": [x / total for x in amean]}}
    if smean:
        out["observation.state"] = {"mean": [x / total for x in smean]}
    return out


def fetch_stats(repo: str):
    s = fetch_json(repo, "stats.json")
    if s and "action" in s:
        return s
    raw = _get(f"https://huggingface.co/datasets/{repo}/resolve/main/meta/episodes_stats.jsonl")
    return aggregate_episodes_stats(raw) if raw else None


def first_episode_video_url(repo: str, info: dict) -> str:
    feats = info.get("features", {}) or {}
    imgs = [k for k in feats if k.startswith("observation.images")]
    vids = [k for k in imgs if feats[k].get("dtype") == "video"] or imgs
    if not vids:
        return ""
    cam = next((k for p in CAM_PRIORITY for k in vids if p in k.lower()), vids[0])
    tmpl = info.get("video_path") or \
        "videos/chunk-{episode_chunk:03d}/{video_key}/episode_{episode_index:06d}.mp4"
    try:
        rel = tmpl.format(episode_chunk=0, episode_index=0, video_key=cam)
    except (KeyError, IndexError, ValueError):
        rel = f"videos/chunk-000/{cam}/episode_000000.mp4"
    return f"https://huggingface.co/datasets/{repo}/resolve/main/{rel}"


def overlap(a_min, a_max, r_min, r_max) -> float:
    """Mean over joints of (intersection width / reference range width)."""
    vals = []
    for i in range(min(len(a_min), len(r_min))):
        lo, hi = max(a_min[i], r_min[i]), min(a_max[i], r_max[i])
        rspan = r_max[i] - r_min[i]
        vals.append(max(0.0, hi - lo) / rspan if rspan else 0.0)
    return round(statistics.mean(vals), 2) if vals else 0.0


def analyze(meta: dict, ref: dict) -> dict:
    row = {c: "" for c in COLUMNS}
    row.update({"id": meta["id"], "url": f"https://huggingface.co/datasets/{meta['id']}",
                "downloads": int(meta["downloads"] or 0), "episodes": int(meta["episodes"] or 0),
                "codebase_version": meta["codebase_version"], "tasks": meta["tasks"][:120]})

    info = fetch_json(meta["id"], "info.json")
    stats = fetch_stats(meta["id"])
    if not info or not stats or "action" not in stats:
        row.update({"flag": "red", "note": "info/stats fetch failed"})
        return row

    row["robot_type"] = str(info.get("robot_type") or "")
    names = [norm_name(n) for n in (info.get("features", {}).get("action", {}).get("names") or [])]
    a = stats["action"]
    dim = len(a["min"])
    names_ok = set(names) == set(STD)
    order_ok = names == STD
    unit = classify_unit(a["min"], a["max"])
    row.update({"dim": dim, "names_ok": "Y" if names_ok else "",
                "order_ok": "Y" if order_ok else "", "unit": unit,
                "action_min": ",".join(f"{x:.0f}" for x in a["min"]),
                "action_max": ",".join(f"{x:.0f}" for x in a["max"])})

    st = stats.get("observation.state")
    off = None
    if st and order_ok and len(st.get("mean", [])) == dim:
        off = round(statistics.mean(abs(a["mean"][i] - st["mean"][i]) for i in range(dim)), 1)
        row["act_state_offset"] = off
    if order_ok and unit == ref["unit"]:
        row["range_overlap"] = overlap(a["min"], a["max"], ref["min"], ref["max"])

    notes = []
    if dim != 6 or not names_ok:
        flag = "red"
        notes.append("dim/name mismatch")
    elif off is not None and off > 15:
        flag = "red"
        notes.append(f"act-state offset {off} deg, strong leader-frame suspicion")
    elif unit != ref["unit"]:
        flag = "yellow"
        notes.append(f"unit {unit} (reference is degree), conversion required")
    elif not order_ok:
        flag = "yellow"
        notes.append("joint order differs, reorder required")
    elif off is not None and off > 6:
        flag = "yellow"
        notes.append(f"act-state offset {off} deg, weak leader-frame suspicion")
    elif off is None:
        flag = "yellow"
        notes.append("no state column, offset unverified")
    else:
        flag = "green"
    row["flag"] = flag
    row["note"] = "; ".join(notes)
    return row


def write_xlsx(rows: list, xlsx_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "action_match"
    hf, hfill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="4472C4")
    lf = Font(color="0563C1", underline="single")
    fills = {"green": PatternFill("solid", fgColor="C6EFCE"),
             "yellow": PatternFill("solid", fgColor="FFEB9C"),
             "red": PatternFill("solid", fgColor="FFC7CE")}
    order = {"green": 0, "yellow": 1, "red": 2}
    rows.sort(key=lambda r: (order.get(r["flag"], 3), -int(r["downloads"] or 0)))

    ws.append(COLUMNS)
    for i in range(1, len(COLUMNS) + 1):
        ws.cell(1, i).font, ws.cell(1, i).fill = hf, hfill
        ws.cell(1, i).alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r.get(c, "") for c in COLUMNS])
        ri = ws.max_row
        ws.cell(ri, 1).fill = fills.get(r["flag"], PatternFill())
        uc = ws.cell(ri, COLUMNS.index("url") + 1)
        uc.hyperlink, uc.value, uc.font = r["url"], "open", lf
        if r.get("video_url"):
            vc = ws.cell(ri, COLUMNS.index("video_url") + 1)
            vc.hyperlink, vc.value, vc.font = r["video_url"], "play", lf
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    widths = {"id": 46, "url": 6, "video_url": 8, "tasks": 40, "note": 34,
              "action_min": 26, "action_max": 26, "codebase_version": 9}
    for i, c in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 10)
    wb.save(xlsx_path)


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--ref", default=os.environ.get("REF", "reference_dataset"),
                    help="reference LeRobot dataset directory (default: $REF)")
    ap.add_argument("--catalog", default="02_so101_catalog.csv", help="stage 1 catalog CSV")
    ap.add_argument("--csv", default="03_action_match.csv", help="output CSV")
    ap.add_argument("--xlsx", default="03_action_match.xlsx", help="output XLSX")
    ap.add_argument("--video", action="store_true",
                    help="attach first-episode preview URLs to green/yellow rows")
    ap.add_argument("--retry", action="store_true",
                    help="re-run only the rows whose metadata fetch failed")
    ap.add_argument("--workers", type=int, default=12, help="fetch workers")
    args = ap.parse_args()

    ref = load_ref(args.ref)
    print(f"[ref] names={ref['names']} unit={ref['unit']}")

    if args.video:
        rows = list(csv.DictReader(open(args.csv, encoding="utf-8")))
        for r in rows:
            r.setdefault("video_url", "")
        tgt = [r for r in rows if r["flag"] in ("green", "yellow")]
        print(f"[video] fetching info.json for {len(tgt)} rows...", flush=True)

        def add(r):
            info = fetch_json(r["id"], "info.json")
            r["video_url"] = first_episode_video_url(r["id"], info) if info else ""
            return r
        with ThreadPoolExecutor(max_workers=8) as ex:
            list(ex.map(add, tgt))
        with open(args.csv, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=COLUMNS)
            w.writeheader()
            w.writerows(rows)
        write_xlsx(rows, args.xlsx)
        got = sum(1 for r in tgt if r["video_url"])
        print(f"[video] {got}/{len(tgt)} preview URLs -> {args.xlsx}")
        return

    if args.retry:
        rows = list(csv.DictReader(open(args.csv, encoding="utf-8")))
        idx = [i for i, r in enumerate(rows) if r["flag"] == "red" and not r["unit"]]
        print(f"[retry] {len(idx)} unresolved rows (workers 4)...", flush=True)
        metas = [{"id": rows[i]["id"], "downloads": rows[i]["downloads"],
                  "episodes": rows[i]["episodes"], "codebase_version": rows[i]["codebase_version"],
                  "tasks": rows[i]["tasks"]} for i in idx]
        with ThreadPoolExecutor(max_workers=4) as ex:
            new = list(ex.map(lambda m: analyze(m, ref), metas))
        for i, nr in zip(idx, new):
            rows[i] = nr
        with open(args.csv, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=COLUMNS)
            w.writeheader()
            w.writerows(rows)
        write_xlsx(rows, args.xlsx)
        cnt = {}
        for r in rows:
            cnt[r["flag"]] = cnt.get(r["flag"], 0) + 1
        still = sum(1 for r in rows if r["flag"] == "red" and not r["unit"])
        print(f"[retry] flag: {cnt} / still unresolved {still}")
        return

    cands = [r for r in csv.DictReader(open(args.catalog, encoding="utf-8"))
             if r["category"] in ("clean", "sim") and int(r["episodes"] or 0) >= 50]
    print(f"[1/2] fetching info+stats for {len(cands)} candidates...", flush=True)

    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        rows = list(ex.map(lambda m: analyze(m, ref), cands))

    with open(args.csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows)
    write_xlsx(rows, args.xlsx)

    cnt = {}
    for r in rows:
        cnt[r["flag"]] = cnt.get(r["flag"], 0) + 1
    units = {}
    for r in rows:
        units[r["unit"] or "?"] = units.get(r["unit"] or "?", 0) + 1
    print(f"[2/2] flag: {cnt}\n  unit distribution: {units}")
    print(f"  -> {args.xlsx} / {args.csv}")


if __name__ == "__main__":
    main()
