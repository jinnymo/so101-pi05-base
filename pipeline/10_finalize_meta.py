# SPDX-License-Identifier: Apache-2.0
"""Stage 6e: repair `info.json` totals from `episodes.jsonl`, then update the report.

Stage 6d finds datasets whose `total_episodes` or `total_frames` disagree with the data
while the data itself is intact. The authoritative counts come from `episodes.jsonl`
(one line per episode, each with its length). Only the metadata is rewritten; parquet and
video files are untouched.

Input:  <root>/external_hf/{confirmed/*/*,excluded/*} and the stage 6a XLSX report.
Output: corrected `meta/info.json` files, and refreshed episodes/frames columns in the XLSX.

Run:
  python 10_finalize_meta.py --root /path/to/workspace --report so101_external_analysis.xlsx
"""

import argparse
import glob
import json
import os


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--root", default=os.environ.get("ROOT", "."),
                    help="workspace root (default: $ROOT)")
    ap.add_argument("--src", default=None, help="dataset directory (default: <root>/external_hf)")
    ap.add_argument("--report", default="so101_external_analysis.xlsx", help="stage 6a XLSX")
    args = ap.parse_args()

    from openpyxl import load_workbook

    src = args.src or os.path.join(args.root, "external_hf")
    real = {}
    fixed = 0
    dirs = glob.glob(f"{src}/confirmed/*/*") + glob.glob(f"{src}/excluded/*")
    for d in dirs:
        if not os.path.isdir(d):
            continue
        ep = f"{d}/meta/episodes.jsonl"
        ij = f"{d}/meta/info.json"
        if not (os.path.exists(ep) and os.path.exists(ij)):
            continue
        eps = [json.loads(line) for line in open(ep, encoding="utf-8") if line.strip()]
        re_ep = len(eps)
        re_fr = sum(int(e.get("length", 0)) for e in eps)
        name = os.path.basename(d)
        real[name] = (re_ep, re_fr)
        info = json.load(open(ij, encoding="utf-8"))
        if info.get("total_episodes") != re_ep or info.get("total_frames") != re_fr:
            info["total_episodes"] = re_ep
            info["total_frames"] = re_fr
            json.dump(info, open(ij, "w", encoding="utf-8"), indent=4)
            fixed += 1
    print(f"info.json repaired: {fixed} datasets")

    wb = load_workbook(args.report)
    ws = wb["datasets"]
    H = [c.value for c in ws[1]]
    di, ei, fi = H.index("dir"), H.index("episodes"), H.index("frames")
    upd = 0
    for row in ws.iter_rows(min_row=2):
        n = row[di].value
        if n in real:
            row[ei].value, row[fi].value = real[n]
            upd += 1
    wb.save(args.report)
    print(f"XLSX episodes/frames updated: {upd} rows")


if __name__ == "__main__":
    main()
